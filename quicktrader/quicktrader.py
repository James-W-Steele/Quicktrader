import serial #pip install pyserial
import time
import win32com.client #pip isntall pypiwin32 
import openpyxl #pip install openpyxl
import os.path
from os import path

def followprogram(program1,wb,trip): #Checks the workbook for discrepancies by following the workbook's path.
    program = program1
    unprinted = {} #cells that were caught by the change command that have yet to be displayed.
    alpha = "abcdefghijklmnopqrstuvwxyz"
    sheetname = ""
    programode = -1
    while programode != 0:
        addmode = 0
        success = []
        fail = []
        current = []
        caught = 0
        programode = 0
        for i in program: #sorts program into logical blocks current, success, and failed
            if i == ["<success>"] and addmode > -1:
                addmode +=1
            if i == ["<failed>"] and addmode < 1:
                addmode -= 1
            if addmode == 0:
                current.append(i)
            if addmode < 0 and i != ["<failed_end>"] and i != ["<failed>"]:
                fail.append(i)
            if addmode > 0 and i != ["<success>"] and i != ["<success_end>"]:
                success.append(i)
            if i == ["<success_end>"] and addmode > -1:
                addmode -=1        
            if i == ["<failed_end>"] and addmode < 1:
                addmode +=1      
        
        for i in current: #follows program
            triggeredcells = []
            if i[0] == "sheet":
                sheet = wb[i[1]]
                sheetname = i[1]
            if i[0] == "refresh":
                if i[1] == "1":
                    trip = refreshvalues(trip)
                    if trip == 1:
                        writetodisplay("Workbook unable to be refreshed")
                    if trip == 0:
                        writetodisplay("Workbook        refreshed")
                if i[1] == "0":
                    trip = refreshvalues(trip)
            if i[0] == "wait":
                time.sleep(int(i[1]))
            if i[0] == "savedata":
                datachecks = []
                miniprogram = []
                for j in program1:
                    if 'check' == j[0]:
                        datachecks.append("["+j[1]+","+j[2]+","+j[3]+"]")
                        miniprogram.append("["+j[1]+","+j[2]+","+j[3]+"]")
                    if 'sheet' == j[0]:
                        miniprogram.append(j[1])
                
                tempdata = {}
                for j in miniprogram:
                    if "[" not in j and "]" not in j and "," not in j:
                        sheet = wb[j]
                    else:
                        location = list(j.split(","))
                        location[0] = location[0][1:]
                        location[-1] = location[-1][:-1]
                        values = get_cell_values(location[0],location[1],location[2])
                        tempdata[j] = values
                savedata(programdatafile,tempdata)
                if i[1] == "1":
                    writetodisplay("Data saved")
                sheet = wb[sheetname]
            if i[0] == "check":
                x = i[1]
                y = i[2]
                z = i[3]
                datakey = "[" + x + "," + y + "," + z + "]"
                savedvalues = retrievedata(programdatafile,datakey)
                currentvalues = get_cell_values(x,y,z)
                cellvals = []
                if not y.isdigit():
                    celval = x+y
                    cellvals.append(celval)
                    y_value = 0
                    z_value = 0
                    for j in y:
                        y_value += ord(j.lower())
                    for j in z:
                        z_value += ord(j.lower())
                    for j in range((z_value - y_value)):
                        if celval[-1].lower() != "z":
                            celval = celval[:-1] + alpha[(alpha.find(celval[-1]))+1]
                        else:
                            bestnotz = -1
                            for k in range(len(celval)):
                                letter = celval[k]
                                if letter.lower() != 'z' and letter.isalpha():
                                    bestnotz = k
                            if bestnotz == -1:
                                newcelval = 'a'
                                for l in celval:
                                    if not l.isdigit():
                                        newcelval += 'a'
                                celval = x+newcelval
                            else:
                                for l in range(len(celval)):
                                    if l == bestnotz:
                                        celval = celval[:l] + alpha[(alpha.find(celval[l]))+1] + celval[l+1:]
                                    if l > bestnotz:
                                        celval = celval[:l] + "a" + celval[l+1:]
                        cellvals.append(celval)
                else:
                    for j in range(int(z)-int(y)+1):
                        cellvals.append(x+str(int(y)+j))
                for j in range(len(savedvalues)):
                    savedvalue = savedvalues[j]
                    currentvalue = currentvalues[j]
                    celval = cellvals[j]
                    if savedvalue != currentvalue and i[4] == 'change':
                        unprinted[celval] = [savedvalue,currentvalue]
                        caught = 1
                    if savedvalue == currentvalue and i[4] == 'same':
                        unprinted[celval] = [savedvalue,currentvalue]
                        caught = 1            
            if i[0] == "output":
                usecells = 0
                for j in i:
                    if j == 'cell_location':
                        usecells = 1
                    test = (j.split("["))
                    if test[0] == 'cell':
                        if 'x_value' in test[1] or 'y_value' in test[1]:
                            usecells = 1                             
                permitoutput = 1
                if usecells == 1 and unprinted == {}:
                    permitoutput = 0
                if permitoutput == 1:
                    if usecells == 0:
                        displaymessage = ""
                        for j in i[1:]:
                            splits = j.split("[")
                            if splits[0] == 'cell':
                                first = ((splits[1])[:-1].split(","))[0]
                                second = ((splits[1])[:-1].split(","))[0]
                                if first.isalpha():
                                    for q in range(int(splits[2].split(",")[0])):
                                        first = next_column(first)
                                else:
                                    for q in range(int((splits[2].split(",")[1])[:-1])):
                                        second = next_column(second)                                                                                     
                                displaymessage += str(get_cell_values(first,second,second)[0])
                            elif j.split("[")[0] == "sheet_change":
                                sheet = wb[(j.split("[")[1])[:-1]]
                            else:
                                displaymessage += j
                        writetodisplay(displaymessage)        
                    else:
                        for k in unprinted.keys():
                            displaymessage = ""
                            sheet = wb[sheetname]
                            currentcell = k
                            currentx = ""
                            currenty = ""
                            alphas = ""
                            digits = ""
                            for m in k:
                                if m.isalpha():
                                    alphas += m
                                else:
                                    digits += m
                            if k[0].isdigit():
                                currentx = digits
                                currenty = alphas
                            else:
                                currentx = alphas
                                currenty = digits
                            for j in i[1:]:
                                splits = j.split("[")
                                if (j.split("["))[0] == 'cell':
                                    first = ((splits[1])[:-1].split(","))[0]
                                    second = ((splits[1])[:-1].split(","))[1]
                                    if first == 'x_value':
                                        first = currentx
                                    if second == 'y_value':
                                        second = currenty
                                    if first.isalpha():
                                        for q in range(int(splits[2].split(",")[0])):
                                            first = next_column(first)
                                        second = str(int(second) + int((splits[2].split(",")[1])[:-1]))
                                    else:
                                        for q in range(int((splits[2].split(",")[1])[:-1])):
                                            second = next_column(second)    
                                        first = str(int(first) + int(int(splits[2].split(",")[0])))
                                    displaymessage += str(get_cell_values(first,second,second)[0])
                                elif j.split("[")[0] == "sheet_change":
                                    sheet = wb[(j.split("[")[1])[:-1]]
                                elif j == "cell_location":
                                    displaymessage += k
                                else:
                                    displaymessage += j
                            writetodisplay(displaymessage)         
                    sheet = wb[sheetname]

        if caught == 0:
            if fail != []:
                programode = -1
                program = fail
            else:
                programode = 0
        else:
            programode = 1
            program = success
    return trip         

def next_column(column):
    alphabet = list('abcdefghijklmnopqrstuvwxyz')
    column = list(column)
    counter = 0
    for i in range(len(column)-1, -1, -1):
        index = alphabet.index(column[i])
        counter += index + 1
        column[i] = alphabet[counter % 26]
    return ''.join(column)   
                
        

def writetodisplay(message): #writes a message to the arduino/display
    ser = serial.Serial(settings['com'], 9600)
    time.sleep(2)
    ser.write(message.encode())
    ser.close()
    print(message)
    time.sleep(int(settings["display message time"]))
    ser = serial.Serial(settings['com'], 9600)
    time.sleep(2)
    ser.close()    
    
def refreshvalues(trip): #refreshes the stock values on the workbook
    available = 1
    try:
            os.rename(settings['excel file'],settings['excel file'])
    except:    
            available = 0
    xlapp = win32com.client.DispatchEx("Excel.Application")
    wb = xlapp.Workbooks.Open(settings['excel file'])
    wb.RefreshAll()   
    if available == 1:
        wb.Save()
        if trip == 1:
            print("The workbook has been closed and refreshes are now working.")
            trip = 0      
    else:
        if trip == 0:
            print("The workbook was not saved because it is currently open. Close the workbook to allow refresh of values.")
            trip += 1       
    wb.Close()
    xlapp.Quit()  
    return trip

def validateprogram(program): #makes sure the program has correct syntax
    errors = []
    successes = 0
    fails = 0
    j = 0
    while j != len(program):
        i = program[j]
        if i[0] == "sheet":
            if len(i) == 2:
                try:
                    wb = openpyxl.load_workbook(settings['excel file'], data_only = True, read_only = True)
                    sheet = wb[i[1]]   
                    wb.close()
                except:
                    errors.append("Sheet access error in line "+str(j+1)+". Make sure you have a sheet named '"+i[1]+"' in the workbook "+settings['excel file'])
            else:
                errors.append("Sheet access field error in line "+str(j+1)+". It looks like the command was given either too few or too many fields. Remember, the sheet command should have two fields")
        elif i[0] == "wipe":
            if len(i) == 2:
                if i[1].lower() in ["no","yes"]:
                    pass
                else:
                    errors.append("Wipe command usage error in line "+str(j+1)+". After using the command, please enter yes or no.")
            else:
                errors.append("Wipe command field error in line "+str(j+1)+". Please use two fields when using the wipe command.")
        elif i[0] == "refresh":
            if len(i) != 2:
                errors.append("Refresh command field error in line "+str(j+1)+". Please use two fields when using the refresh command.")
            else:
                if i[1] not in ["1","0"]:
                    errors.append("Refresh command syntax error in line "+str(j+1)+". Please enter a 1 or a 0 for the second field.")            
        elif i[0] == "check":
            if len(i) == 5:
                if i[1].isdigit():
                    if i[2].isalpha() and i[3].isalpha():
                        pass
                    else:
                        errors.append("Check command syntax error in line "+str(j+1)+". When referencing a row, please enter a column range with the next two values.")
                elif i[1].isalpha():
                    if i[2].isdigit() and i[3].isdigit():
                        pass
                    else:
                        errors.append("Check command syntax error in line "+str(j+1)+". When referencing a column, please enter a row range with the next two values.")   
                else:
                    errors.append("Check command syntax error in line "+str(j+1)+". Please reference a row or column with the second field")
                if i[4] not in ["same","change"]:
                    errors.append("Check command syntax error in line "+str(j+1)+". Please use either 'change' or 'same' in the fifth field.")
            else:
                errors.append("Check command field error in line "+str(j+1)+". Please use five fields when using the check command.")
        elif i[0] == "savedata":
            if len(i) != 2:
                errors.append("Savedata command field error in line "+str(j+1)+". Please input two fields when using the savedata command.")   
            else:
                if i[1] not in ["0","1"]:
                    errors.append("Savedata command syntax error in line "+str(j+1)+". Please input either a 1 or a 0 for the second field.")   
        elif i[0] == "<success>":
            if len(i) != 1:
                errors.append("Success command field error in line "+str(j+1)+". Please only use one field when using the <success> command.")      
            successes += 1
        elif i[0] == "<success_end>":
            if len(i) != 1:
                errors.append("Success end command field error in line "+str(j+1)+". Please only use one field when using the <success_end> command.")  
            successes -= 1
        elif i[0] == "<failed>":
            if len(i) != 1:
                errors.append("Failed command field error in line "+str(j+1)+". Please only use one field when using the <failed> command.")       
            fails += 1
        elif i[0] == "<failed_end>":
            if len(i) != 1:
                errors.append("Success command field error in line "+str(j+1)+". Please only use one field when using the <failed_end> command.")         
            fails -= 1
        elif i[0] == "output":
            pass
            #if len(i) != 3:
                #errors.append("Output command field error in line "+str(j+1)+". Please use 3 fields when using the output command. Consult the manual if need be.")    
            #else:
                #if i[1] not in ['message','value','location']:
                    #errors.append("Output command usage error in line "+str(j+1)+". Please use one of the following when using the output command: message, value, or location.")
        elif i[0] == "wait":
            if len(i) != 2:
                errors.append("Wait command field error in line "+str(j+1)+". Please use 2 fields when using the wait command.")    
            else:
                if not i[1].isdigit():
                    errors.append("Wait command syntax error in line "+str(j+1)+". Please specify a number of seconds to wait in the second field.")
        elif i[0] == "<data_start>":
            j = len(program)-1
        else:
            errors.append("The command in line "+str(j+1)+"was not valid. Please rewrite that line.")
        j += 1
    if fails > 0:
        errors.append("Fail function usage error. There are more instances of the <failed> command than the <failed_end> command. Please make sure to add <failed_end> after finishing your function.")
    if fails < 0:
        errors.append("Fail function usage error. There are more instances of the <failed_end> command than the <failed> command. Please make sure to mark when your function starts.")
    if successes > 0:
        errors.append("Success function usage error. There are more instances of the <success> command than the <success_end> command. Please make sure to add <success_end> after finishing your function.")
    if successes < 0:
        errors.append("Success function usage error. There are more instances of the <success_end> command than the <success> command. Please make sure to mark when your function starts.")
    return errors

def get_cell_values(x,y,z): #gets values from the worksheet.
    x_value = 0
    y_value = 0
    z_value = 0
    resultvalues = []
    for i in x:
        if i.isdigit():
            x_value = int(x)
        else:
            x_value += ord(i.lower()) - 96
    for i in y:
        if i.isdigit():
            y_value = int(y)
        else:
            y_value += ord(i.lower()) - 96
    for i in z:
        if i.isdigit():
            z_value = int(z)+1
        else:
            z_value += ord(i.lower()) - 95
    if x.isdigit():
        for k in range (y_value,z_value):
            resultvalues.append(sheet.cell(row=x_value,column=k).value)          
    else:
        for k in range (y_value,z_value):
            resultvalues.append(sheet.cell(row=k,column=x_value).value)       
    if y_value == z_value:
        if x.isdigit():
            resultvalues.append(sheet.cell(row=x_value,column=y_value).value)
        else:
            resultvalues.append(sheet.cell(row=y_value,column=x_value).value)
    return resultvalues

data = {}

def savedata(file,data): #saves data to the data file
    savefile = open("data/"+file,"w")
    for i in data.keys():
        savefile.write(str(i)+"\n")
        for j in data[i]:
            savefile.write(str(j)+"\n")
        savefile.write("[end]\n")
    savefile.close()
    
def retrievedata(file,data): #gets data from already saved data file
    savefile = open("data/"+file,"r")
    savedata = savefile.readlines()
    savefile.close()
    for i in range(len(savedata)):
        savedata[i] = savedata[i][:-1]
    mode = 0
    result = []   
    for i in savedata:
        if mode == 1 and i == '[end]':
            mode = 0
        if mode == 1:
            result.append(i)
        if mode == 0:
            if i == data:
                mode = 1
    return result
                

def preparedata(program,wb): #checks that the data file exists, creates it if it does not, and wipes existing data files if wipe is set to 1
    wipe = 0
    anticounter = 0
    for j in range(len(program)):
        i = program[j-anticounter]
        if i[0] == "wipe":
            if i[1].lower() == 'yes':
                wipe = 1
            program = program[:j] + program[j+1:]
            anticounter += 1
    datachecks = []
    miniprogram = []
    for i in program:
        if 'check' == i[0]:
            datachecks.append("["+i[1]+","+i[2]+","+i[3]+"]")
            miniprogram.append("["+i[1]+","+i[2]+","+i[3]+"]")
        if 'sheet' == i[0]:
            miniprogram.append(i[1])
    
    tempdata = {}
    for i in miniprogram:
        if "[" not in i and "]" not in i and "," not in i:
            sheet = wb[i]
        else:
            location = list(i.split(","))
            location[0] = location[0][1:]
            location[-1] = location[-1][:-1]
            values = get_cell_values(location[0],location[1],location[2])
            tempdata[i] = values
            
    filexist = 0
    try:
        test = open("data/"+programdatafile,"r+")
        filedata = test.readlines()
        for i in range(len(filedata)):
            filedata[i] = filedata[i][:-1]
        test.close()
        filexist = 1
    except:
        filedata = ""
        
    
    checked = []
    for i in filedata:
        if "[" in i and "]" in i and "," in i and "end" not in i:
            checked.append(i)
    if filexist == 1:
        corrupt = 0
        for i in datachecks:
            if i not in checked:
                corrupt = 1
        if corrupt == 1:
            wb.close()
            x = input("The data file looks like it does not have the correct data associated with this program.\nTerminate the program to save the existing data. \nIf you press enter, the file will be erased the file for you and rewritten with current data:")
            wb=openpyxl.load_workbook(settings['excel file'])
            sheet=wb[program[0][1]]
            wipe = 1
    if filexist == 1:
        if wipe == 1:
            savedata(programdatafile,tempdata)
            print("Program data wiped and refreshed.")
        else:
            print("Existing program data loaded.")
    else:
        savedata(programdatafile,tempdata)
        print("Data file created and current data loaded.")
    
        
            
        
    

trip = 0
settingsfileopen = open('settings.txt', 'r')
settingsfile = settingsfileopen.readlines()
settingsfileopen.close()
settings = {}
for i in range(len(settingsfile)): #derives information from settings file
    if settingsfile[i] != settingsfile[-1]:
        settingsfile[i] = settingsfile[i][:-1]
    j = settingsfile[i].split(":")
    result = ""
    for i in j:
        if i == j[0]:
            pass
        elif i == j[1]:
            result = result + i
        else:
            result = result + ":" + i
    settings.update({j[0]:result})

for i in settings.keys(): #if the user has elected to input fresh settings data each time the program runs, this will ask them for it.
    if settings[i] == "ask":
        settings[i] = input("Please put the value for "+i+" here:")

validated = 0
while validated != len(settings.keys()): #This whole while loop validates the information gathered from the settings file.
    validated = 0
    for i in settings.keys():
        if i == "program":
            try:
                tried = open(settings[i],'r')
                tried.close()
                validated += 1
            except:
                settings[i] = input("The program you entered was not able to be opened. Please enter the correct program title and corresponding location here:")
        elif i == "com":
            try:
                ser = serial.Serial(settings[i], 9600)
                ser.close()
                validated += 1
            except:
                settings[i] = input("The com port you entered was not able to be communicated with. Please check if the device is plugged in and that your com port is correct. \nPlease enter the revised com port here:")    #remember to delete the comment out.
        elif i == "excel file":
            try:
                wb = openpyxl.load_workbook(settings[i], data_only = True, read_only = True) 
                wb.close()
                validated += 1
            except:
                settings[i] = input("The excel workbook you entered was not able to be opened. Please enter the correct workbook title and corresponding location here:")
        elif i == "cooldown":
            if not settings[i].isdigit():
                settings[i] = input("The cooldown you entered was not a number. Please enter a cooldown containing only digits here:")
            else:
                validated += 1
        elif i == "display message time":
            if not settings[i].isdigit():
                settings[i] = input("The display message time you entered was not a number. Please enter a cooldown containing only digits here:")
            else:
                validated += 1            
        else:
            x = input("It looks like "+i+" is in the settings file. This is not a setting. The program will proceed anyway, but please delete the excess line. press enter to acknowledge:")
            validated += 1
        if "program" not in settings.keys() or "com" not in settings.keys() or "excel file" not in settings.keys() or "cooldown" not in settings.keys():
            x = input("It looks like a setting is missing from the settings file. Remember, the settings are: com, excel file, cooldown, and program. \nPlease terminate the program and correct the settings file.\nConsult the manual if need be.\nPress enter to acknowledge:")
print("Settings file validated.")
programdatafile = ((((settings['program']).split(".txt"))[0])+"_data.txt")
programfile = open(settings['program'],'r') #gathers information from program file.
programraw = programfile.readlines()
programfile.close()
program = []
for i in range(len(programraw)): #Sorts through the lines and preps the program for validation.
    if programraw[i] != programraw[-1]:
        programraw[i] = programraw[i][:-1]+" "
    j = []
    parenthesis = 0
    currentword = ""
    for l in range(len(programraw[i])):
        k = programraw[i][l]
        if k == " " and parenthesis == 0:
            if currentword != '':            
                j.append(currentword)
            currentword = ""
        elif k == "(":
            parenthesis = 1
        elif k == ")":
            parenthesis = 0
            if l == len(programraw[i])-1:
                j.append(currentword)
        elif k == " " and parenthesis == 1:
            currentword += k
        else:
            currentword += k
    if j != []:
        program.append(j)
print("Program data sorted.")
    
errors = validateprogram(program)
if len(errors) > 0:
    print("\n\n")
    for i in errors:
        print(i)
    writetodisplay(str(len(errors))+" error(s) found in program.")
    x = input("Please terminate the program and resolve the errors.")
else:
    print("Program validated.")


wb=openpyxl.load_workbook(settings['excel file'])

sheet=wb[program[0][1]]
preparedata(program,wb)

x = 0
writetodisplay("Program running")
while x!= 1:
    trip = followprogram(program,wb,trip)
    print("program finished")
    time.sleep(int(settings['cooldown']))

wb.close()
#wb = openpyxl.load_workbook(settings['excel file'], data_only = True, read_only = True) 
#sheet = wb["Sheet1"]
  
#c = sheet['F5'] 

#print(c.value)
#trip = refreshvalues(trip)
#trip = refreshvalues(trip)
